import openpyxl as op;
import sys;

path = sys.argv[1]

wb = op.load_workbook(path)
ws = wb.active

cell = ws.cell(row=1, column=1)

print(ws.max_row)
print(ws.max_column)

for column in range(1, ws.max_column):
    for row in range(1, ws.max_row):
        cell = ws.cell(row=row, column=column)

        if cell.value:
            print(cell.value)
